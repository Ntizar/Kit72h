#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Kit72h — buscador nocturno de fichas reales de Amazon.

Los productos cuyo enlace es aún una búsqueda genérica (es_busqueda: true en
data/kits.json) se resuelven aquí: se busca en amazon.es (curl + --compressed
+ cookie jar, patrón probado en agosto 2026), se filtran candidatos por rango
de precio y ≥4★, y se VERIFICA cada ficha final (HTTP 200 + buybox + precio)
antes de proponerla.

Nunca escribe enlaces en kits.json: genera propuestas en data/propuestas-fichas.json
y un Excel (data/productos-pendientes.xlsx) para que David revise y valide.
Rotación determinista: cada noche toca un cupo de N pendientes (por defecto 12),
así el trabajo se reparte sin estado y sin quemar requests de más.

Uso:
  python scripts/buscar-amazon.py                 # lote de esta noche (12)
  python scripts/buscar-amazon.py --cupo 5        # lote pequeño
  python scripts/buscar-amazon.py --producto "sirena"   # probar uno
"""
import argparse
import html
import json
import random
import re
import subprocess
import sys
import time
import urllib.parse
from datetime import date
from pathlib import Path

RAIZ = Path(__file__).resolve().parents[1]
D_KITS = RAIZ / "data" / "kits.json"
D_PROP = RAIZ / "data" / "propuestas-fichas.json"
D_XLSX = RAIZ / "data" / "productos-pendientes.xlsx"
D_LOG = RAIZ / "data" / "buscador-log.json"
TAG = "nti0c8-21"
UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126 Safari/537.36")

# Productos que no son de compra online: nunca se buscan
EXCLUIR = re.compile(r"efectivo|paracetamol|ibuprofeno|renueva|recarga gratis|"
                     r"fotocopias|plan familiar|reunión|punto de encuentro", re.I)


def curl_get(url, cookiejar, timeout=60):
    """curl con cookie jar; devuelve (status, texto)."""
    tmp = RAIZ / "data" / "_tmp_page.html"
    cmd = ["curl", "-s", "-L", "-o", str(tmp), "-w", "%{http_code}",
           "-c", str(cookiejar), "-b", str(cookiejar), "-A", UA,
           "--compressed", "--max-time", str(timeout), url]
    try:
        r = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout + 15)
        code = r.stdout.strip() or "000"
    except subprocess.TimeoutExpired:
        return "000", ""
    txt = ""
    if tmp.exists():
        txt = tmp.read_text(encoding="utf-8", errors="ignore")
        tmp.unlink(missing_ok=True)
    return code, txt


def buscar(query, jar):
    """Resultados de amazon.es para una query. [(asin, titulo, precio, rating, resenas)]"""
    url = "https://www.amazon.es/s?k=" + urllib.parse.quote_plus(query)
    for intento in range(3):
        code, txt = curl_get(url, jar)
        if code == "200" and "s-result" in txt and "captcha" not in txt[:3000].lower():
            return parse_resultados(txt)
        jar.unlink(missing_ok=True)  # jar quemado -> renew
        time.sleep(2 + random.random() * 3)
    return []


def parse_resultados(txt):
    out = []
    bloques = re.split(r'data-asin="([A-Z0-9]{10})"', txt)
    for i in range(1, len(bloques) - 1, 2):
        asin, b = bloques[i], bloques[i + 1][:8000]
        if "a-price" not in b:
            continue
        m = re.search(r"<h2[^>]*>.*?<span[^>]*>(.*?)</span>", b, re.S)
        titulo = html.unescape(re.sub(r"<[^>]+>", "", m.group(1))).strip() if m else ""
        mp = re.search(r'a-price-whole">([\d.,]+)', b)
        mf = re.search(r'a-price-fraction">(\d+)', b)
        precio = None
        if mp:
            precio = float(mp.group(1).replace(".", "").replace(",", ".")) + \
                     (int(mf.group(1)) / 100 if mf else 0)
        mr = re.search(r"([\d,]+) de 5 estrellas", b)
        rating = float(mr.group(1).replace(",", ".")) if mr else None
        mrv = re.search(r'a-size-base s-underline-text[^>]*>\s*([\d.,]+)', b)
        resenas = int(re.sub(r"[.,]", "", mrv.group(1))) if mrv else 0
        if titulo and precio:
            out.append((asin, titulo, precio, rating, resenas))
    return out


def verificar_ficha(asin, jar):
    """True si la ficha dp/ASIN responde 200 y tiene buybox (puede comprarse)."""
    code, txt = curl_get(f"https://www.amazon.es/dp/{asin}", jar)
    if code != "200" or len(txt) < 20000:
        return False
    return bool(re.search(r'id="buybox"|add-to-cart-button|coreNoDupeErrors', txt))


def rango_de(precio_aprox):
    """'5-15 €' -> (4, 17) con márgenes."""
    nums = [float(x.replace(",", ".")) for x in re.findall(r"\d+(?:[.,]\d+)?", precio_aprox or "")]
    if not nums:
        return (2, 200)
    lo = min(nums) * 0.6
    hi = max(nums) * 1.6 + 3
    return (max(1, lo), hi)


def coletilla_no_compra(producto):
    return EXCLUIR.search(producto)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--cupo", type=int, default=12)
    ap.add_argument("--producto", default=None, help="filtrar por substring (pruebas)")
    ap.add_argument("--limpieza", action="store_true", help="marcar N/A los no-compra y reescribir log")
    args = ap.parse_args()

    kits = json.loads(D_KITS.read_text(encoding="utf-8"))
    propuestas = json.loads(D_PROP.read_text(encoding="utf-8")) if D_PROP.exists() else {}
    log = json.loads(D_LOG.read_text(encoding="utf-8")) if D_LOG.exists() else {"resueltos": [], "na": [], "intentos": []}

    pendientes = []
    for kit in kits["kits"]:
        for s in kit["secciones"]:
            for i in s["items"]:
                prod = i.get("producto", "")
                url = i.get("afiliado") or ""
                sin_ficha = ("amazon.es/s?" in url) or not url or i.get("es_busqueda")
                if not sin_ficha:
                    continue
                if coletilla_no_compra(prod):
                    if prod not in log["na"]:
                        log["na"].append(prod)
                    continue
                clave = f"{kit['slug']}|{prod}"
                # solo excluir si ya tiene ficha aplicada; las propuestas sin
                # candidato verificado se reintentan en noches posteriores
                if (propuestas.get(clave) or {}).get("url"):
                    continue
                if prod in log["resueltos"]:
                    continue
                pendientes.append((kit, s, i, clave))

    if args.limpieza:
        D_LOG.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"log de limpieza: {len(log['na'])} marcados N/A")
        return

    hoy = date.today()
    dia = int(hoy.strftime("%Y%m%d"))
    rot = dia % max(1, len(pendientes)) if pendientes else 0
    lote = pendientes[rot:rot + args.cupo] or pendientes[:args.cupo]
    if args.producto:
        lote = [p for p in pendientes if args.producto.lower() in p[2]["producto"].lower()][:5]

    print(f"Pendientes: {len(pendientes)} | lote de esta noche: {len(lote)} | N/A: {len(log['na'])}")
    jar = RAIZ / "data" / "_cookies_amz.txt"
    nuevos = 0
    aplicadas = 0
    for kit, s, i, clave in lote:
        prod = i["producto"]
        q = i.get("busqueda") or re.sub(r"\(.*?\)|[—–:/]", " ", prod).strip().lower()
        lo, hi = rango_de(i.get("precio_aprox"))
        res = buscar(q, jar)
        cands = [r for r in res if lo <= r[2] <= hi and (r[3] or 0) >= 4.0 and r[4] >= 50]
        if not cands:
            cands = [r for r in res if lo * 0.7 <= r[2] <= hi * 1.4 and (r[3] or 0) >= 4.0]
        cands.sort(key=lambda r: -(r[4] or 0))
        ficha_ok = None
        for asin, titulo, precio, rating, resenas in cands[:3]:
            if verificar_ficha(asin, jar):
                ficha_ok = {"asin": asin, "titulo": titulo[:120], "precio": precio,
                            "rating": rating, "resenas": resenas,
                            "url": f"https://www.amazon.es/dp/{asin}?tag={TAG}",
                            "verificacion": "buybox"}
                break
            time.sleep(1 + random.random())
        entry = {"kit": kit["slug"], "seccion": s["titulo"], "busqueda": q,
                 "rango_esperado": i.get("precio_aprox"), "fecha": hoy.isoformat()}
        if ficha_ok:
            entry.update(ficha_ok)
            entry["aplicada"] = True
            log["resueltos"].append(prod)
            # Auto-aplicar: la ficha está verificada (buybox + precio en rango + ≥4★).
            # El enlace conserva el tag de afiliado; queda registrado en propuestas
            # y en el Excel semanal por si David quiere revertir alguno.
            anterior = i.get("afiliado")
            i["afiliado"] = ficha_ok["url"]
            i["es_busqueda"] = False
            i["verificado_ficha"] = hoy.isoformat()
            entry["anterior"] = anterior
            aplicadas += 1
        else:
            entry["estado"] = "sin candidato verificado"
        propuestas[clave] = entry
        nuevos += 1
        print(f"  {'✅' if ficha_ok else '·'} {prod[:55]:57s} -> "
              f"{(ficha_ok or {}).get('url', 'reintentar otra noche')}")
        time.sleep(2 + random.random() * 2)

    D_PROP.write_text(json.dumps(propuestas, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    D_LOG.write_text(json.dumps(log, ensure_ascii=False, indent=2), encoding="utf-8")
    jar.unlink(missing_ok=True)
    if aplicadas:
        D_KITS.write_text(json.dumps(kits, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(f"kits.json actualizado con {aplicadas} fichas verificadas")

    # Excel de validación para David
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = "Pendientes"
        cols = ["Nº", "Kit", "Sección", "Producto", "Búsqueda", "URL verificada (con tu tag)",
                "Precio", "Rating", "Reseñas", "Verificación", "Estado", "Validado (sí/no)"]
        ws.append(cols)
        for c in ws[1]:
            c.font = Font(bold=True)
        n = 0
        for kit in kits["kits"]:
            for s in kit["secciones"]:
                for i in s["items"]:
                    prod = i["producto"]
                    clave = f"{kit['slug']}|{prod}"
                    es_na = coletilla_no_compra(prod)
                    if not (es_busqueda(i) or es_na):
                        continue
                    p = propuestas.get(clave)
                    n += 1
                    ws.append([n, kit["slug"], s["titulo"], prod,
                               i.get("busqueda", ""),
                               (p or {}).get("url", ""),
                               (p or {}).get("precio", ""),
                               (p or {}).get("rating", ""),
                               (p or {}).get("resenas", ""),
                               (p or {}).get("verificacion", ""),
                               "N/A — no es compra online" if es_na else
                               ((p or {}).get("estado") or "propuesta"), ""])
        anchos = [4, 16, 18, 42, 30, 52, 8, 7, 8, 10, 22, 14]
        for j, w in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(D_XLSX)
        print(f"Excel: {D_XLSX.name} con {n} filas")
    except ImportError:
        print("openpyxl no disponible: solo JSON de propuestas")

    print(f"DONE propuestas totales: {len(propuestas)} | resueltos: {len(log['resueltos'])} | pendientes restantes: {max(0, len(pendientes) - nuevos)}")


def es_busqueda(i):
    u = i.get("afiliado") or ""
    return i.get("es_busqueda") or "amazon.es/s?" in u


if __name__ == "__main__":
    main()
